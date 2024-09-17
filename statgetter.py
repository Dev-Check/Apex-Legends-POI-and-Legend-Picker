import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox

maps = {
    "storm": ["barometer", "beast", "camp", "cascade", "cenote", "checkpoint", "coast", "command", "echo", "launch", "lightning", "mill", "north", "pylon", "station", "storm", "wall", "zeus"], 
    "moon": ["array", "atmostation", "base", "bionomics", "cliff", "core", "cultivation", "divide", "foundry", "gardens", "gulch", "labs", "port", "production", "quarantine", "solar", "terraformer", "underpas", "wharf"],
    "district": ["bank", "blossom", "boardwalk", "dam", "galleria", "hall", "heights", "lotus", "market", "point", "resort", "shipyard", "square", "stadium", "viaduct"],  
    "kings": ["airbase", "artillery", "basin", "bunker", "cage", "caustic", "containmment", "crash", "gauntlet", "hydrodam", "labs", "lake", "map", "market", "pit", "relic", "repulsor", "rig", "runoff", "swamps"],
    "worlds": ["camp", "climatizer", "countdown", "dome", "epicenter", "fissure", "fragment", "geyser", "harvester", "launch", "maude", "overlook", "siphon", "skyhook", "staging", "thermal", "trials", "tree"],
    "olympus": ["array", "bonsai", "cannon", "carrier", "clinic", "depot", "docks", "estates", "gardens", "grid", "hammond", "hydroponics", "icarus", "oly", "phase", "rift", "towers", "turbine"]
}

legends = ["wattson", "wraith", "loba", "alter", "ash", "ballistic", "bangalore", "bloodhound", "catalyst", "caustic", "conduit", "crypto", "fuse", "gibralter", "horizon", "lifeline", "mad maggie", "mirage", "newcastle", "octane", "pathfinder", "rampart", "revenant", "seer", "valkyrie", "vantage"]
numbers = [str(i) for i in range(0, 21)]

# Define the sheet name directly in the code
sheetname = "storm"  # Update this to the sheet name you want to use

# Function to load the Excel file
def loadfile(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name)
        return workbook
    except FileNotFoundError:
        print(f"Error: The file {file_name} does not exist.")
        exit(1)
    except PermissionError:
        print(f"Error: Permission denied for file {file_name}.")
        exit(1)

# Function to log data into the correct sheet
def logdata(workbook, sheetname, legend, poi, kills, placement, file_name):
    if sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        # Append the data to the sheet
        sheet.append([legend, poi, kills, placement])
        try:
            workbook.save(file_name)  # Save using the correct file_name variable
            print(f"Data logged in {sheetname}: {legend}, {poi}, {kills} kills, placed {placement}")
        except PermissionError:
            print(f"Error: Permission denied while saving the file {file_name}.")
    else:
        print(f"Error: {sheetname} sheet does not exist. Available sheets: {workbook.sheetnames}")

def updatedropdown(*args):
    selected_map = getmap.get()
    pois = maps[selected_map]
    poidropdown['values'] = pois
    getpoi.set(pois[0])  # Set the default POI to the first one in the list

def submitdata():
    currmap = getmap.get()
    currlegend = getlegend.get()
    poi = getpoi.get()
    kills = getkills.get()
    placement = getplacement.get()

    if not (currmap and currlegend and poi and kills and placement):
        messagebox.showerror("Error", "All fields are required!")
        return
    logdata(workbook, currmap, currlegend, poi, int(kills), int(placement), file_name)

file_name = "apexavg.xlsx"  # Use your existing Excel file
workbook = loadfile(file_name)

# Create the main window
root = tk.Tk()
root.title("Apex Legends Game Logger")

# Map selection
getmap = tk.StringVar(value=list(maps.keys())[0])
tk.Label(root, text="Select Map:").grid(row=0, column=0, padx=10, pady=10)
mapdropdown = ttk.Combobox(root, textvariable=getmap, values=list(maps.keys()), state="readonly")
mapdropdown.grid(row=0, column=1)

# Legend selection
getlegend = tk.StringVar(value=legends[0])
tk.Label(root, text="Select Legend:").grid(row=1, column=0, padx=10, pady=10)
legenddropdown = ttk.Combobox(root, textvariable=getlegend, values=legends, state="readonly")
legenddropdown.grid(row=1, column=1)

# POI dropdown
getpoi = tk.StringVar(value=maps[list(maps.keys())[0]][0])
tk.Label(root, text="Select POI:").grid(row=2, column=0, padx=10, pady=10)
poidropdown = ttk.Combobox(root, textvariable=getpoi, values=maps[list(maps.keys())[0]], state="readonly")
poidropdown.grid(row=2, column=1)

# Kills dropdown
getkills = tk.StringVar(value=numbers[0])
tk.Label(root, text="Select Kills:").grid(row=3, column=0, padx=10, pady=10)
killsdropdown = ttk.Combobox(root, textvariable=getkills, values=numbers, state="readonly")
killsdropdown.grid(row=3, column=1)

# Placement dropdown
getplacement = tk.StringVar(value=numbers[0])
tk.Label(root, text="Select Placement:").grid(row=4, column=0, padx=10, pady=10)
placementdropdown = ttk.Combobox(root, textvariable=getplacement, values=numbers, state="readonly")
placementdropdown.grid(row=4, column=1)

# Submit button
submitbutton = tk.Button(root, text="Submit", command=submitdata)
submitbutton.grid(row=5, column=0, columnspan=2, pady=20)

# Update POI dropdown when a new map is selected
getmap.trace_add("write", updatedropdown)

# Start the GUI
root.mainloop()
"""
def main():
    file_name = "apexavg.xlsx"  # Use your existing Excel file
    workbook = loadfile(file_name)

    # Print existing sheet names for debugging
    print("Current Sheet:\n", sheetname)
    
        

    while True:
        # Ask the user for inputs
        legend = input("Enter the legend you used: ").strip()
        poi = input("Enter the POI where you landed: ").strip()
        kills = int(input("Enter the number of kills: "))
        placement = int(input("Enter your placement: "))

        # Log the data to the predefined sheet
        logdata(workbook, legend, poi, kills, placement, file_name)

        # Ask if the user wants to continue logging
        cont = input("Do you want to log another game? (y/n): ")
        if cont.lower() != 'y':
            break

if __name__ == "__main__":
    main()
"""